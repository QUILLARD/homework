import openpyxl as xl

wb_01 = xl.load_workbook("1111.xlsx")
wb_02 = xl.load_workbook("2222.xlsx")
wb_03 = xl.load_workbook("3333.xlsx")
ws1 = wb_01.active
ws2 = wb_02.active
ws3 = wb_03.active

pages = []
for ws in [ws1, ws2, ws3]:
    print(f"Страница: {ws}")
    for row in ws.iter_rows(values_only=True):
        print(f"Колонка: {row}")
        pages.append(row)

data = sorted(pages, reverse=True)

wb = xl.Workbook()
ws = wb.active

for row in data:
    ws.append(row)

for row in ws.iter_rows():
    for cell in row:
        cell.font = xl.styles.Font(name="Arial Black")
        cell.border = xl.styles.Border(left=xl.styles.Side(border_style="double", color="0000FF00"),
                                       right=xl.styles.Side(border_style="double", color="0000FF00"),
                                       top=xl.styles.Side(border_style="double", color="0000FF00"),
                                       bottom=xl.styles.Side(border_style="double", color="0000FF00"))

wb.save("result.xlsx")
